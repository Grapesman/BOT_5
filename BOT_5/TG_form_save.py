from datetime import datetime
from openpyxl import load_workbook
def how_much_string(file_path):
    a = []
    i = 1
    book = load_workbook(filename = file_path)
    sheet = book['Каталог статей']
    while sheet['A' + str(i)].value is not None:
        a.append(i)
        i+=1
    return len(a)

async def function_TG_form_save(save_path, file_path, data1, data2, data3, data4, data5):
    today = datetime.today()
    today_date = today.strftime("%d.%m.%Y")
    stroki = how_much_string(file_path)
    book = load_workbook(save_path)
    sheet = book['Каталог статей']
    sheet['A' + str(stroki + 1)].value = str(stroki + 2)
    sheet['B' + str(stroki + 1)].value = data1
    sheet['D' + str(stroki + 1)].value = data2
    sheet['F' + str(stroki + 1)].value = data3
    sheet['H' + str(stroki + 1)].value = data4
    sheet['J' + str(stroki + 1)].value = data5
    sheet['C' + str(stroki + 1)].value = datetime.strptime(today_date, "%d.%m.%Y")
    print("Данные успешно записаны в файл")
    # Сохранение изменений в файле
    book.save(file_path)